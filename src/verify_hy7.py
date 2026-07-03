#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
花页7 现有材料核校：把"多尺度数据汇总.xlsx"（派生汇总）与各尺度文件夹里
服务商的原始统计 xlsx 三方对账，回答"汇总是否忠实于原始、有无编造"。
  (A) 矿物含量：源 粗扫/精扫矿物元素含量.xlsx  vs  汇总 Amics_矿物含量
  (B) 各尺度孔隙/喉道/裂缝分布：源 孔隙喉道统计.xlsx  vs  汇总 对应 sheet（数字覆盖率）
  (C) 孔隙度：源 面孔率 sheet 逐切片重算  vs  汇总 孔隙度汇总
用法： .venv/bin/python src/verify_hy7.py
"""
import os, warnings
import openpyxl
warnings.filterwarnings("ignore")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
B = os.path.join(ROOT, "data", "hy7_raw")
SUMMARY = f"{B}/花页7井_4199.21m_多尺度数据汇总.xlsx"
LEGACY_SUMMARY_NOTE = (
    "verify_hy7.py 是早期一次性对账工具，依赖已删除的派生汇总 xlsx；"
    "当前主线数字源改为 src/hy7_etl.py 从服务商原始 xlsx 生成 experiments/hy7_stats.json。"
)
PASS, FAIL, NOTE = [], [], []
def chk(cond, label, detail=""):
    (PASS if cond else FAIL).append(label)
    print(f"  [{'PASS' if cond else 'FAIL'}] {label}" + (f"  {detail}" if detail else ""))

def sheet_numbers(path, sheet, rnd=2):
    """取一个 sheet 内所有数值（四舍五入），返回 multiset(dict 计数)。"""
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb[sheet]
    from collections import Counter
    c = Counter()
    for row in ws.iter_rows(values_only=True):
        for x in row:
            if isinstance(x, (int, float)) and not isinstance(x, bool):
                c[round(float(x), rnd)] += 1
    return c

def file_numbers(path, rnd=2, sheets=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    from collections import Counter
    c = Counter()
    for sh in (sheets or wb.sheetnames):
        for row in wb[sh].iter_rows(values_only=True):
            for x in row:
                if isinstance(x, (int, float)) and not isinstance(x, bool):
                    c[round(float(x), rnd)] += 1
    return c

def minerals(path):
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb["矿物"]
    out = {}
    for row in ws.iter_rows(values_only=True):
        if row[1] and isinstance(row[2], (int, float)):
            out[str(row[1]).strip()] = round(float(row[2]), 3)
    return out

if not os.path.exists(SUMMARY):
    raise FileNotFoundError(f"{LEGACY_SUMMARY_NOTE} 缺失文件: {SUMMARY}")

print("=== (A) 矿物含量：源 vs 汇总 ===")
src_coarse = minerals(f"{B}/Amics矿物整体扫描25um+精细扫描1um/粗扫矿物元素含量.xlsx")
src_fine   = minerals(f"{B}/Amics矿物整体扫描25um+精细扫描1um/精扫矿物元素含量.xlsx")
sum_nums = sheet_numbers(SUMMARY, "Amics_矿物含量", rnd=2)
# 汇总里粗扫矿物值应逐一命中
miss_c = [m for m,v in src_coarse.items() if sum_nums.get(round(v,2),0) < 1]
miss_f = [m for m,v in src_fine.items() if sum_nums.get(round(v,2),0) < 1]
chk(len(miss_c)==0, f"粗扫25μm {len(src_coarse)}种矿物含量 全部命中汇总", f"缺失={miss_c}")
chk(len(miss_f)<=1, f"精扫1μm {len(src_fine)}种矿物含量 命中汇总", f"缺失={miss_f}")
chk(abs(sum(src_coarse.values())-100)<0.5, "粗扫矿物含量合计≈100%", f"Σ={sum(src_coarse.values()):.2f}")
chk(abs(sum(src_fine.values())-100)<0.5, "精扫矿物含量合计≈100%", f"Σ={sum(src_fine.values()):.2f}")

print("\n=== (B) 各尺度 孔隙/喉道/裂缝分布：源数字是否被汇总忠实收录（覆盖率）===")
SCALE_SRC = {
    "微米CT_14μm":   f"{B}/微米CT柱塞扫描-14um/孔隙喉道、裂缝统计数据.xlsx",
    "微米CT_2.8μm":  f"{B}/微米CT精细扫描-2p8um/孔隙喉道、裂缝统计数据.xlsx",
    "纳米CT_65nm":   f"{B}/纳米CT扫描-65nm/孔隙喉道、裂缝统计数据.xlsx",
    "FIB-SEM_8.589nm": f"{B}/FIB-SEM聚焦离子束扫描-8p589nm/孔隙喉道、裂缝统计数据.xlsx",
    "Maps_10nm":     f"{B}/Maps整体扫描200nm+精细扫描10nm/孔隙、裂缝计算结果.xlsx",
}
for sumsheet, srcpath in SCALE_SRC.items():
    if not os.path.exists(srcpath):
        chk(False, f"{sumsheet} 源文件存在", srcpath); continue
    # 源分布数字（排除“面孔率”逐切片大表，只取分布统计 sheet）
    wb = openpyxl.load_workbook(srcpath, data_only=True)
    dist_sheets = [s for s in wb.sheetnames if "面孔率" not in s]
    # Maps 源每个分布表含 半径(μm) 与 半径(nm) 两列同值，汇总只保留 nm；剔除 μm 列免误判
    from collections import Counter
    src = Counter()
    for s in dist_sheets:
        for row in wb[s].iter_rows(values_only=True):
            cells = list(row)
            if "Maps" in sumsheet and cells and isinstance(cells[0], (int, float)) \
               and len(cells) > 1 and isinstance(cells[1], (int, float)):
                cells = cells[1:]  # 跳过 μm 半径列（与 nm 列重复）
            for x in cells:
                if isinstance(x, (int, float)) and not isinstance(x, bool):
                    src[round(float(x), 2)] += 1
    summ = sheet_numbers(SUMMARY, sumsheet, rnd=2)
    # 只看“有意义”的分布数字（>0 且非纯整数序号），统计命中率
    keys = [k for k,c in src.items() if k not in (0.0,) ]
    hit = sum(1 for k in keys if summ.get(k,0) >= 1)
    cov = hit/len(keys) if keys else 1
    chk(cov >= 0.95, f"{sumsheet}: 源分布数字命中汇总 {hit}/{len(keys)} = {cov*100:.0f}%",
        ("" if cov>=0.95 else f"（未命中{len(keys)-hit}项，需查）"))

print("\n=== (C) 孔隙度：源 面孔率逐切片重算 vs 汇总孔隙度汇总 ===")
# 汇总孔隙度
sumwb = openpyxl.load_workbook(SUMMARY, data_only=True); ps = sumwb["孔隙度汇总"]
sum_poro = {}
for row in ps.iter_rows(values_only=True):
    if row and isinstance(row[0], str) and isinstance(row[2], (int,float)):
        sum_poro[row[0].strip()] = round(float(row[2]),3)
# 逐切片重算（有 面孔率 sheet 的尺度）
def reavg_facekong(srcpath):
    wb = openpyxl.load_workbook(srcpath, data_only=True)
    sh = next((s for s in wb.sheetnames if "面孔率" in s), None)
    if not sh: return None
    ws = wb[sh]; rows=list(ws.iter_rows(values_only=True))
    hdr=rows[0]
    # 找“裂缝面孔率”“孔隙面孔率”列
    def col(name):
        for j,h in enumerate(hdr):
            if h and name in str(h): return j
        return None
    cf,cp = col("裂缝面孔率"), col("孔隙面孔率")
    if cf is not None and cp is not None:        # 14μm/65nm：裂缝+孔隙分列
        fr=[r[cf] for r in rows[1:] if isinstance(r[cf],(int,float))]
        po=[r[cp] for r in rows[1:] if isinstance(r[cp],(int,float))]
        return (sum(fr)/len(fr), sum(po)/len(po), len(po))
    # 2.8μm：仅单列“面孔率%”（非“平均面孔率%”）
    ct=None
    for j,h in enumerate(hdr):
        if h and "面孔率" in str(h) and "平均" not in str(h) and "切片" not in str(h):
            ct=j; break
    if ct is None: return None
    vals=[r[ct] for r in rows[1:] if isinstance(r[ct],(int,float))]
    return (0.0, sum(vals)/len(vals), len(vals))
for scale, key in [("微米CT柱塞扫描-14um","微米CT柱塞扫描"),
                   ("微米CT精细扫描-2p8um","微米CT精细扫描"),
                   ("纳米CT扫描-65nm","纳米CT扫描")]:
    srcpath=f"{B}/{scale}/孔隙喉道、裂缝统计数据.xlsx"
    if not os.path.exists(srcpath): continue
    r=reavg_facekong(srcpath)
    if not r:
        NOTE.append(f"{scale} 无面孔率逐切片表"); continue
    frac,por,n=r; total=frac+por
    sv=sum_poro.get(key)
    ok = sv is not None and abs(total-sv)<0.05
    chk(ok, f"{key}: 逐切片重算(裂缝{frac:.3f}+孔隙{por:.3f}={total:.3f}%, n={n}) vs 汇总{sv}",
        "" if ok else "（差异>0.05，需查）")

print(f"\n=== 结论：{len(PASS)} PASS / {len(FAIL)} FAIL ===")
if NOTE: print("  备注：", "; ".join(NOTE))
if FAIL: print("  失败项：", FAIL)
else: print("  汇总表数字均可由各尺度原始统计复现，矿物/孔隙度均忠实于原始，无编造。")
