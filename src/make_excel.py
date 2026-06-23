#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成 Excel 数据分析手册：deliverables/GJ5-15_数据分析手册.xlsx
数据来源：experiments/master_stats.json + data/warehouse.db（均由 etl_build_warehouse.py 产出）。
统计量为管线计算值（已验证），全量连续曲线随附以便复算。
用法： .venv/bin/python src/make_excel.py
"""
import os, json
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ST = json.load(open(os.path.join(ROOT, "experiments", "master_stats.json"), encoding="utf-8"))
DB = os.path.join(ROOT, "data", "warehouse.db")
FIG = os.path.join(ROOT, "experiments", "figures")
OUT = os.path.join(ROOT, "deliverables", "GJ5-15_数据分析手册.xlsx")

FONT = "Arial"
NAVY = "1F3864"; BLUE = "2E5496"; LIGHT = "D6E4F0"; GREY = "F2F2F2"
HDR = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE = Font(name=FONT, bold=True, size=16, color=NAVY)
SUB = Font(name=FONT, bold=True, size=12, color=BLUE)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

CURVE_ORDER = ST["curve_order"]
META = ST["curve_meta"]
SEGS = ST["segments"]


def hdr_row(ws, row, headers, start=1, fill=BLUE):
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=start + j, value=h)
        c.font = HDR; c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = CEN; c.border = BORDER


def box(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


# ---------------- 0. 导读 ----------------
def sheet_readme(wb):
    ws = wb.active; ws.title = "导读"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "GJ5-15 数字井筒 · 数据分析手册"; ws["A1"].font = TITLE
    ws["A2"] = "Generative Digital Wellbore — Data Reference Workbook"
    ws["A2"].font = Font(name=FONT, size=10, italic=True, color="808080")
    rows = [
        ("", ""),
        ("井号 / 样品", ST["well"]),
        ("岩心井段", f"{ST['interval_m'][0]} – {ST['interval_m'][1]} m（共 {ST['interval_length_m']} m）"),
        ("连续曲线采样点", f"{ST['n_samples']:,} 点 / 条，垂向分辨率约 {ST['vertical_resolution_mm']} mm"),
        ("CT 扫描规格", f"像素 {ST['ct_spec']['pixel_um']} μm，{ST['ct_spec']['image']}，{ST['ct_spec']['bit_depth']}，{ST['ct_spec']['format']}"),
        ("深度分段", "  /  ".join(SEGS)),
        ("", ""),
        ("本手册内容", ""),
        ("  数据字典", "每条曲线/参数的中英文、缩写、单位、物理含义、取值范围、来源文件"),
        ("  分段统计", "5 个深度段 × 9 条曲线的均值/最小/最大/标准差（核心表）"),
        ("  全井统计", "全井段每条曲线的分布统计（均值、中位数、分位、极值）"),
        ("  相关性矩阵", "9 条曲线两两 Pearson 相关（逐层）"),
        ("  渗透率", "分段水平/垂直渗透率 Kh、Kv 及各向异性 Kh/Kv"),
        ("  孔隙直径分布", "孔隙等效直径分级：按数量占比 vs 按体积贡献"),
        ("  储层评价", "沿井深 a/b/c/d 综合评价分段及厚度汇总"),
        ("  连续曲线(原始)", "全量逐层曲线 17,982 行，供绘图与复算"),
        ("  图表", "井筒综合道、分段对比、孔径分布、相关性、渗透率、储层评价"),
        ("", ""),
        ("数据来源", "外置盘原始统计（CT 反演逐层 txt/xlsx）→ src/etl_build_warehouse.py → DuckDB + master_stats.json"),
        ("数值说明", "统计量为 ETL 管线计算值；已与原始逐片数据精确复现（src/verify_integrity.py，15 项核验全过）"),
        ("分段均值核验", "本手册各属性分段【均值】与服务商汇总表自报值逐项吻合（最大偏差 0.02）"),
        ("min/max 口径", "本手册 min/max 为逐片原始极值；服务商汇总表按深度聚合后报，极差更窄——两者均值一致，非错误"),
        ("生成日期", "2026-06-22"),
    ]
    r = 4
    for k, v in rows:
        a = ws.cell(row=r, column=1, value=k)
        b = ws.cell(row=r, column=2, value=v)
        a.font = BOLD if k and not k.startswith("  ") else BODY
        if k in ("本手册内容",): a.font = SUB
        b.font = BODY; b.alignment = LEFT
        r += 1
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 95


# ---------------- 1. 数据字典 ----------------
def sheet_dict(wb):
    ws = wb.create_sheet("数据字典"); ws.sheet_view.showGridLines = False
    ws["A1"] = "数据字典 Data Dictionary"; ws["A1"].font = SUB
    hdr_row(ws, 3, ["中文名", "英文 key", "缩写", "单位", "全井范围", "物理含义 / 说明"])
    abbr = {"matrix_porosity":"KX","total_porosity":"-","fracture_porosity":"LF","total_vug_frac":"KDF",
            "vug_fill_rate":"KDFCT","oil_content":"HYL","oil_saturation":"HYBHD","shale_content":"NZ","diagenetic_min":"CYKW"}
    desc = {
        "matrix_porosity":"岩石基质（粒间/粒内）孔隙体积占比，储集空间主体",
        "total_porosity":"基质孔隙 + 孔洞缝总孔隙体积占比",
        "fracture_porosity":"裂缝贡献的孔隙体积占比，反映裂缝发育程度",
        "total_vug_frac":"孔、洞、缝综合体积比率",
        "vug_fill_rate":"孔洞缝被矿物/沥青等充填的比率，越高有效连通越差",
        "oil_content":"含油孔隙体积 / 岩心体积，CT 反演含油率",
        "oil_saturation":"含油孔隙体积 / 总孔隙体积，含油饱和度",
        "shale_content":"泥质（黏土）体积占比 Vsh，越高物性通常越差",
        "diagenetic_min":"成岩自生矿物体积占比，本井极低(均值<0.05%)",
    }
    r = 4
    for k in CURVE_ORDER:
        o = ST["overall"][k]
        vals = [META[k]["zh"], k, abbr[k], META[k]["unit"], f"{o['min']} ~ {o['max']}", desc[k]]
        for j, v in enumerate(vals):
            c = ws.cell(row=r, column=1 + j, value=v); c.font = BODY; c.border = BORDER
            c.alignment = LEFT if j in (1, 5) else CEN
        r += 1
    # 附：储层评价、渗透率、孔径说明
    extra = [
        ("储层综合评价", "grade", "-", "a/b/c/d", "a 优 → d 差", "沿井深分段定性评价，见「储层评价」表"),
        ("水平渗透率", "kh_mD", "Kh", "mD", "6.4 ~ 87.8", "平行层理方向渗透率，分段值"),
        ("垂直渗透率", "kv_mD", "Kv", "mD", "2.9 ~ 13.3", "垂直层理方向渗透率，分段值"),
        ("渗透率各向异性", "kh_kv", "Kh/Kv", "-", "1.95 ~ 6.59", "Kh/Kv 比值，反映层理/裂缝定向性"),
        ("孔隙等效直径", "pore_dia", "-", "μm", "50 ~ >1000", "孔隙分级，见「孔隙直径分布」表"),
    ]
    for k in extra:
        for j, v in enumerate(k):
            c = ws.cell(row=r, column=1 + j, value=v); c.font = BODY; c.border = BORDER
            c.fill = PatternFill("solid", fgColor=GREY)
            c.alignment = LEFT if j in (1, 5) else CEN
        r += 1
    widths = [16, 18, 8, 8, 16, 56]
    for i, w in enumerate(widths): ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A4"


# ---------------- 2. 分段统计（核心） ----------------
def sheet_segment(wb):
    ws = wb.create_sheet("分段统计"); ws.sheet_view.showGridLines = False
    ws["A1"] = "各深度段统计汇总（单位 %）"; ws["A1"].font = SUB
    ws["A2"] = ("每个深度段 × 每条曲线：均值 / 最小 / 最大 / 标准差。均值与服务商汇总表逐项吻合；"
                "min/max 为逐片原始极值（服务商按深度聚合后报，极差更窄）。")
    ws["A2"].font = Font(name=FONT, size=9, color="808080")
    # 表头：属性 + 各段
    hdr_row(ws, 4, ["曲线", "统计量"] + SEGS)
    r = 5
    mean_rows = []
    for k in CURVE_ORDER:
        zh = META[k]["zh"]
        for stat, label in [("mean", "均值"), ("min", "最小"), ("max", "最大"), ("std", "标准差")]:
            ws.cell(row=r, column=1, value=zh).font = BOLD if stat == "mean" else BODY
            ws.cell(row=r, column=2, value=label).font = BODY
            for j, seg in enumerate(SEGS):
                v = ST["per_segment"][seg][k][stat]
                c = ws.cell(row=r, column=3 + j, value=v)
                c.font = BOLD if stat == "mean" else BODY
                c.number_format = "0.000"; c.alignment = CEN
            if stat == "mean": mean_rows.append(r)
            r += 1
    box(ws, 4, 1, r - 1, 2 + len(SEGS))
    # 给均值行加色阶（按行：跨段对比）
    for mr in mean_rows:
        rng = f"{get_column_letter(3)}{mr}:{get_column_letter(2+len(SEGS))}{mr}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B"))
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=1)
    ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 9
    for j in range(len(SEGS)): ws.column_dimensions[get_column_letter(3 + j)].width = 15
    ws.freeze_panes = "C5"
    # 渗透率 + 评价小结附在右侧下方
    base = r + 2
    ws.cell(row=base, column=1, value="分段渗透率 (mD) 与储层评价").font = SUB
    hdr_row(ws, base + 1, ["深度段", "Kh", "Kv", "Kh/Kv"])
    for i, p in enumerate(ST["permeability"]):
        vals = [p["segment_label"], p["kh_mD"], p["kv_mD"], p["kh_kv_ratio"]]
        for j, v in enumerate(vals):
            c = ws.cell(row=base + 2 + i, column=1 + j, value=v); c.font = BODY; c.border = BORDER
            c.alignment = CEN
            if j >= 1: c.number_format = "0.0"


# ---------------- 3. 全井统计 ----------------
def sheet_overall(wb):
    ws = wb.create_sheet("全井统计"); ws.sheet_view.showGridLines = False
    ws["A1"] = "全井段分布统计（3439.13–3443.52 m，n=17,982，单位 %）"; ws["A1"].font = SUB
    cols = ["曲线", "均值", "标准差", "最小", "P25", "中位数", "P75", "最大"]
    keys = ["mean", "std", "min", "p25", "median", "p75", "max"]
    hdr_row(ws, 3, cols)
    r = 4
    for k in CURVE_ORDER:
        o = ST["overall"][k]
        ws.cell(row=r, column=1, value=META[k]["zh"]).font = BOLD
        for j, sk in enumerate(keys):
            c = ws.cell(row=r, column=2 + j, value=o[sk]); c.font = BODY
            c.number_format = "0.000"; c.alignment = CEN
        r += 1
    box(ws, 3, 1, r - 1, len(cols))
    ws.column_dimensions["A"].width = 14
    for j in range(1, len(cols)): ws.column_dimensions[get_column_letter(1 + j)].width = 11
    ws.freeze_panes = "B4"


# ---------------- 4. 相关性 ----------------
def sheet_corr(wb):
    ws = wb.create_sheet("相关性矩阵"); ws.sheet_view.showGridLines = False
    ws["A1"] = "曲线相关性矩阵（逐层 Pearson）"; ws["A1"].font = SUB
    zh = [META[k]["zh"] for k in CURVE_ORDER]
    hdr_row(ws, 3, [""] + zh)
    for i, ki in enumerate(CURVE_ORDER):
        ws.cell(row=4 + i, column=1, value=META[ki]["zh"]).font = HDR
        ws.cell(row=4 + i, column=1).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(row=4 + i, column=1).alignment = CEN; ws.cell(row=4 + i, column=1).border = BORDER
        for j, kj in enumerate(CURVE_ORDER):
            v = ST["correlation"][ki][kj]
            c = ws.cell(row=4 + i, column=2 + j, value=round(v, 2)); c.font = BODY
            c.number_format = "0.00"; c.alignment = CEN; c.border = BORDER
    n = len(CURVE_ORDER)
    rng = f"B4:{get_column_letter(1+n)}{3+n}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num", start_value=-1, start_color="4472C4",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="num", end_value=1, end_color="C00000"))
    ws.column_dimensions["A"].width = 14
    for j in range(n): ws.column_dimensions[get_column_letter(2 + j)].width = 11
    ws.freeze_panes = "B4"


# ---------------- 5. 孔隙直径 ----------------
def sheet_pore(wb):
    ws = wb.create_sheet("孔隙直径分布"); ws.sheet_view.showGridLines = False
    ws["A1"] = "孔隙等效直径分布"; ws["A1"].font = SUB
    hdr_row(ws, 3, ["直径区间 (μm)", "按数量占比", "按体积贡献"])
    for i, p in enumerate(ST["pore_diameter"]):
        ws.cell(row=4 + i, column=1, value=p["diameter_um"]).font = BODY
        for j, key in enumerate(["count_fraction", "volume_fraction"]):
            c = ws.cell(row=4 + i, column=2 + j, value=p[key]); c.font = BODY
            c.number_format = "0.0%"; c.alignment = CEN
    box(ws, 3, 1, 3 + len(ST["pore_diameter"]), 3)
    ws.cell(row=4 + len(ST["pore_diameter"]) + 1, column=1,
            value="提示：小孔（50–100μm）数量占绝对多数，但大孔贡献了主要孔隙体积——储集空间由少数大孔主导。").font = Font(name=FONT, size=9, italic=True, color="808080")
    for i, w in enumerate([16, 14, 14]): ws.column_dimensions[get_column_letter(i + 1)].width = w


# ---------------- 6. 储层评价 ----------------
def sheet_grade(wb):
    ws = wb.create_sheet("储层评价"); ws.sheet_view.showGridLines = False
    ws["A1"] = "储层综合评价分段（沿井深）"; ws["A1"].font = SUB
    hdr_row(ws, 3, ["顶深 (m)", "底深 (m)", "厚度 (m)", "评价等级"])
    cmap = {"a": "63BE7B", "b": "B7E1A1", "c": "FFE699", "d": "F8696B"}
    r = 4
    for it in ST["reservoir_grade_intervals"]:
        th = round(it["end_depth"] - it["start_depth"], 3)
        vals = [it["start_depth"], it["end_depth"], th, it["grade"].upper()]
        for j, v in enumerate(vals):
            c = ws.cell(row=r, column=1 + j, value=v); c.font = BODY; c.border = BORDER; c.alignment = CEN
            if j < 3: c.number_format = "0.000"
        ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=cmap.get(it["grade"], "FFFFFF"))
        r += 1
    # 厚度汇总
    r += 1
    ws.cell(row=r, column=1, value="各等级累计厚度 (m)").font = SUB; r += 1
    hdr_row(ws, r, ["等级", "厚度 (m)", "占比"])
    total = sum(ST["reservoir_grade_thickness"].values())
    r += 1
    names = {"a":"a 优", "b":"b 良", "c":"c 中", "d":"d 差"}
    for g in ["a", "b", "c", "d"]:
        th = ST["reservoir_grade_thickness"].get(g, 0)
        ws.cell(row=r, column=1, value=names[g]).font = BODY; ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=cmap[g])
        c2 = ws.cell(row=r, column=2, value=round(th, 3)); c2.font = BODY; c2.border = BORDER; c2.alignment = CEN; c2.number_format="0.000"
        c3 = ws.cell(row=r, column=3, value=round(th/total, 4) if total else 0); c3.font = BODY; c3.border = BORDER; c3.alignment = CEN; c3.number_format="0.0%"
        r += 1
    for i, w in enumerate([12, 12, 10]): ws.column_dimensions[get_column_letter(i + 1)].width = w


# ---------------- 7. 连续曲线原始 ----------------
def sheet_curves(wb):
    con = duckdb.connect(DB, read_only=True)
    df = con.execute(f"SELECT depth_m, segment_label, {', '.join(CURVE_ORDER)} FROM curves").df()
    con.close()
    ws = wb.create_sheet("连续曲线(原始)")
    headers = ["深度(m)", "深度段"] + [META[k]["zh"] for k in CURVE_ORDER]
    hdr_row(ws, 1, headers)
    for _, row in df.iterrows():
        ws.append([round(row["depth_m"], 4), row["segment_label"]] +
                  [round(float(row[k]), 3) for k in CURVE_ORDER])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 11; ws.column_dimensions["B"].width = 17
    for j in range(len(CURVE_ORDER)): ws.column_dimensions[get_column_letter(3 + j)].width = 12


# ---------------- 8. 图表 ----------------
def sheet_charts(wb):
    ws = wb.create_sheet("图表"); ws.sheet_view.showGridLines = False
    ws["A1"] = "分析图表（由 src/make_figures.py 生成）"; ws["A1"].font = SUB
    imgs = [("02_segment_bars.png", "A3"), ("03_pore_diameter.png", "A45"),
            ("04_correlation.png", "A78"), ("06_reservoir_grade.png", "L78")]
    for fn, anchor in imgs:
        p = os.path.join(FIG, fn)
        if os.path.exists(p):
            im = XLImage(p)
            scale = 760 / im.width if "02_" in fn else 560 / im.width
            im.width = int(im.width * scale); im.height = int(im.height * scale)
            ws.add_image(im, anchor)


def main():
    wb = Workbook()
    sheet_readme(wb)
    sheet_dict(wb)
    sheet_segment(wb)
    sheet_overall(wb)
    sheet_corr(wb)
    sheet_pore(wb)
    sheet_grade(wb)
    sheet_curves(wb)
    sheet_charts(wb)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(">> 已生成", os.path.relpath(OUT, ROOT))
    print("   sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
