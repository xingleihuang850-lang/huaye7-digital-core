#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""花页7 Excel 多尺度数据手册 → deliverables/花页7/花页7_多尺度数据手册.xlsx"""
import os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

ROOT=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ST=json.load(open(os.path.join(ROOT,"experiments","hy7_stats.json"),encoding="utf-8"))
FIG=os.path.join(ROOT,"experiments","hy7_figures")
OUT=os.path.join(ROOT,"deliverables","花页7","花页7_多尺度数据手册.xlsx")
FONT="Arial"; NAVY="10233A"; BLUE="1C7293"; GREY="F2F2F2"
HDR=Font(name=FONT,bold=True,color="FFFFFF",size=11)
SUB=Font(name=FONT,bold=True,size=12,color="1C7293")
TITLE=Font(name=FONT,bold=True,size=16,color=NAVY)
BODY=Font(name=FONT,size=10); BOLD=Font(name=FONT,size=10,bold=True)
THIN=Side(style="thin",color="BFBFBF"); BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

def hdr(ws,row,headers,start=1):
    for j,h in enumerate(headers):
        c=ws.cell(row=row,column=start+j,value=h); c.font=HDR
        c.fill=PatternFill("solid",fgColor=BLUE); c.alignment=CEN; c.border=BORDER
def box(ws,r1,c1,r2,c2):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1): ws.cell(row=r,column=c).border=BORDER

def sheet_readme(wb):
    ws=wb.active; ws.title="导读"; ws.sheet_view.showGridLines=False
    ws["A1"]="花页7井 多尺度数字岩心 · 数据手册"; ws["A1"].font=TITLE
    ws["A2"]="HuaYe-7 Multi-scale Digital Rock — Data Reference"; ws["A2"].font=Font(name=FONT,size=10,italic=True,color="808080")
    g=ST["lithology_groups"]
    rows=[("",""),
        ("井 / 深度 / 岩性","花页7 / 4199.21 m / 页岩（钙质长英质）"),
        ("数据来源",ST["source"]),
        ("成像尺度","6 个：25μm Amics → 14μm/2.8μm 微米CT → 65nm 纳米CT → 200/10nm Maps SEM → 8.589nm FIB-SEM"),
        ("多尺度总孔隙度","1.56 / 7.18 / 1.93 / 0.41 %（14μm/2.8μm/65nm/Maps，非单调=尺度效应/REV）"),
        ("矿物组成","长英质 {}% + 碳酸盐 {}% + 黏土 {}% + 黄铁矿 {}%".format(g['长英质'],g['碳酸盐'],g['黏土'],g['黄铁矿'])),
        ("FIB-SEM","孔隙 {:,} / 喉道 {:,}；有机质占比 {}%（非孔隙度）".format(ST['fib_counts']['pores'],ST['fib_counts']['throats'],ST['fib_organic_pct'])),
        ("",""),
        ("本手册内容",""),
        ("  尺度总览","6 个成像尺度的模态/分辨率/维度/内容/孔隙度"),
        ("  多尺度孔隙度","各尺度孔隙度 + Maps 有机/无机分类"),
        ("  矿物组成","Amics 粗扫25μm + 精扫1μm + 岩性归组"),
        ("  孔隙半径分布","各尺度孔隙半径(体积/个数占比)"),
        ("  喉道与配位","各尺度喉道半径/长度 + 配位数(连通性)"),
        ("  Maps有机无机","有机/无机 孔隙与裂缝半径分布"),
        ("  图表","多尺度孔隙度、矿物、孔径、连通性、尺度阶梯"),
        ("",""),
        ("数据可信度",ST["verified"]),
        ("生成日期","2026-06-22"),
    ]
    r=4
    for k,v in rows:
        a=ws.cell(row=r,column=1,value=k); b=ws.cell(row=r,column=2,value=v)
        a.font=BOLD if (k and not k.startswith("  ")) else BODY
        if k=="本手册内容": a.font=SUB
        b.font=BODY; b.alignment=LEFT; r+=1
    ws.column_dimensions["A"].width=18; ws.column_dimensions["B"].width=92

def sheet_scales(wb):
    ws=wb.create_sheet("尺度总览"); ws.sheet_view.showGridLines=False
    ws["A1"]="六个成像尺度总览"; ws["A1"].font=SUB
    hdr(ws,3,["尺度","分辨率","模态","数据维度","主要内容","总孔隙度%"])
    r=4
    for s in ST["scales"]:
        vals=[s["name"],s["res"],s["modality"],s["dim"],s["content"],s["porosity"] if s["porosity"] else "—"]
        for j,v in enumerate(vals):
            c=ws.cell(row=r,column=1+j,value=v); c.font=BODY; c.border=BORDER
            c.alignment=LEFT if j in (3,4) else CEN
        r+=1
    for i,w in enumerate([14,12,16,22,28,9]): ws.column_dimensions[get_column_letter(i+1)].width=w

def sheet_porosity(wb):
    ws=wb.create_sheet("多尺度孔隙度"); ws.sheet_view.showGridLines=False
    ws["A1"]="多尺度孔隙度（沿用原始报告“总孔隙度”；FIB-SEM 的 1.52% 是有机质占比，非孔隙度）"; ws["A1"].font=SUB
    hdr(ws,3,["尺度","分辨率","指标","数值%","数据来源","备注"])
    r=4
    for p in ST["porosity_summary"]:
        vals=[p["scale"],p["res"],p["metric"],p["porosity"],p.get("source"),p.get("note")]
        for j,v in enumerate(vals):
            c=ws.cell(row=r,column=1+j,value=v); c.font=BODY; c.border=BORDER
            c.alignment=LEFT if j in (4,5) else CEN
            if j==3: c.number_format="0.000"
        r+=1
    for i,w in enumerate([14,9,16,9,24,30]): ws.column_dimensions[get_column_letter(i+1)].width=w
    # Maps 分类
    r+=2; ws.cell(row=r,column=1,value="Maps SEM(10nm) 孔隙分类").font=SUB; r+=1
    hdr(ws,r,["孔隙类型","面积率%","面占比%","平均半径nm"]); r+=1
    for c in ST["maps_classification"]:
        vals=[c["type"],c["area_pct"],c["share_pct"],c["radius_nm"]]
        for j,v in enumerate(vals):
            cc=ws.cell(row=r,column=1+j,value=v); cc.font=BODY; cc.border=BORDER; cc.alignment=CEN
        r+=1

def sheet_minerals(wb):
    ws=wb.create_sheet("矿物组成"); ws.sheet_view.showGridLines=False
    ws["A1"]="矿物组成（Amics）"; ws["A1"].font=SUB
    hdr(ws,3,["矿物","粗扫25μm %","精扫1μm %"])
    cm={d["mineral"]:d["pct"] for d in ST["minerals"]["coarse_25um"]}
    fm={d["mineral"]:d["pct"] for d in ST["minerals"]["fine_1um"]}
    names=list(dict.fromkeys(list(cm)+list(fm)))
    names.sort(key=lambda n:-cm.get(n,0))
    r=4
    for n in names:
        for j,v in enumerate([n,cm.get(n,"—"),fm.get(n,"—")]):
            c=ws.cell(row=r,column=1+j,value=v); c.font=BODY; c.border=BORDER
            c.alignment=LEFT if j==0 else CEN
            if j>0 and isinstance(v,(int,float)): c.number_format="0.00"
        r+=1
    box(ws,3,1,r-1,3)
    r+=1; ws.cell(row=r,column=1,value="岩性归组（粗扫）").font=SUB; r+=1
    hdr(ws,r,["类别","含量%"]); r+=1
    for k,v in ST["lithology_groups"].items():
        for j,x in enumerate([k,v]):
            c=ws.cell(row=r,column=1+j,value=x); c.font=BODY; c.border=BORDER; c.alignment=CEN
            if j==1: c.number_format="0.00"
        r+=1
    for i,w in enumerate([14,13,13]): ws.column_dimensions[get_column_letter(i+1)].width=w

def _stack(ws,start_row,title_map,getrows,headers):
    r=start_row
    for key,label in title_map:
        ws.cell(row=r,column=1,value=label).font=SUB; r+=1
        hdr(ws,r,headers); r+=1
        for row in getrows(key):
            for j,v in enumerate(row):
                c=ws.cell(row=r,column=1+j,value=v); c.font=BODY; c.border=BORDER; c.alignment=CEN
                if j>0 and isinstance(v,(int,float)): c.number_format="0.000"
            r+=1
        r+=1
    return r

def sheet_pore_radius(wb):
    ws=wb.create_sheet("孔隙半径分布"); ws.sheet_view.showGridLines=False
    ws["A1"]="各尺度孔隙半径分布"; ws["A1"].font=SUB
    tm=[("ct14","微米CT 14μm（半径 μm）"),("ct28","微米CT 2.8μm（半径 μm）"),
        ("nano65","纳米CT 65nm（半径 μm）"),("fib","FIB-SEM 8.589nm（半径 nm）")]
    _stack(ws,3,tm,lambda k:[[d["bin"],d["v1"],d["v2"]] for d in ST["scale_data"][k]["pore_radius"]],
           ["半径区间","体积占比%","个数占比%"])
    for i,w in enumerate([16,13,13]): ws.column_dimensions[get_column_letter(i+1)].width=w

def sheet_throat(wb):
    ws=wb.create_sheet("喉道与配位"); ws.sheet_view.showGridLines=False
    ws["A1"]="各尺度喉道半径与孔隙配位数（连通性）"; ws["A1"].font=SUB
    tm=[("ct14","微米CT 14μm"),("ct28","微米CT 2.8μm"),("nano65","纳米CT 65nm"),("fib","FIB-SEM")]
    r=3
    for k,lab in tm:
        ws.cell(row=r,column=1,value=lab+" 喉道半径").font=SUB; r+=1
        hdr(ws,r,["喉道半径区间","体积占比%","个数占比%"]); r+=1
        for d in ST["scale_data"][k].get("throat_radius",[]):
            for j,v in enumerate([d["bin"],d["v1"],d["v2"]]):
                c=ws.cell(row=r,column=1+j,value=v); c.font=BODY; c.border=BORDER; c.alignment=CEN
                if j>0 and isinstance(v,(int,float)): c.number_format="0.000"
            r+=1
        # 配位数
        ws.cell(row=r,column=5,value=lab+" 配位数").font=SUB
        cr=r+1; ws.cell(row=cr,column=5,value="配位数").font=HDR
        ws.cell(row=cr,column=5).fill=PatternFill("solid",fgColor=BLUE)
        ws.cell(row=cr,column=6,value="个数占比%").font=HDR; ws.cell(row=cr,column=6).fill=PatternFill("solid",fgColor=BLUE)
        cr+=1
        for d in ST["scale_data"][k].get("coordination",[]):
            ws.cell(row=cr,column=5,value=d["bin"]).font=BODY; ws.cell(row=cr,column=5).border=BORDER; ws.cell(row=cr,column=5).alignment=CEN
            ws.cell(row=cr,column=6,value=d["v2"]).font=BODY; ws.cell(row=cr,column=6).border=BORDER; ws.cell(row=cr,column=6).number_format="0.000"; ws.cell(row=cr,column=6).alignment=CEN
            cr+=1
        r=max(r,cr)+1
    for i,w in enumerate([16,13,13,2,12,12]): ws.column_dimensions[get_column_letter(i+1)].width=w

def sheet_maps(wb):
    ws=wb.create_sheet("Maps有机无机"); ws.sheet_view.showGridLines=False
    ws["A1"]="Maps SEM(10nm) 有机/无机 孔隙·裂缝半径分布"; ws["A1"].font=SUB
    tm=[(k,k) for k in ST["maps_dist"]]
    _stack(ws,3,tm,lambda k:[[d["r_nm"],d["pct"]] for d in ST["maps_dist"][k]],
           ["半径/开度 nm","分布占比%"])
    for i,w in enumerate([16,14]): ws.column_dimensions[get_column_letter(i+1)].width=w

def sheet_charts(wb):
    ws=wb.create_sheet("图表"); ws.sheet_view.showGridLines=False
    ws["A1"]="分析图表（src/hy7_figures.py 生成）"; ws["A1"].font=SUB
    imgs=[("01_porosity.png","A3"),("02_minerals.png","A30"),("03_pore_radius.png","A58"),
          ("04_coordination.png","A92"),("06_ladder.png","A118")]
    for fn,anchor in imgs:
        p=os.path.join(FIG,fn)
        if os.path.exists(p):
            im=XLImage(p); scale=720/im.width
            im.width=int(im.width*scale); im.height=int(im.height*scale); ws.add_image(im,anchor)

def main():
    wb=Workbook()
    sheet_readme(wb); sheet_scales(wb); sheet_porosity(wb); sheet_minerals(wb)
    sheet_pore_radius(wb); sheet_throat(wb); sheet_maps(wb); sheet_charts(wb)
    os.makedirs(os.path.dirname(OUT),exist_ok=True); wb.save(OUT)
    print(">> 已生成",os.path.relpath(OUT,ROOT),"| sheets:",wb.sheetnames)

if __name__=="__main__": main()
