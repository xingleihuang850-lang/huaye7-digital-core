#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
花页7 多尺度数据 ETL：从已核校的服务商原始 xlsx 抽取多尺度孔隙度、矿物、
各尺度孔喉/配位/裂缝分布 → experiments/hy7_stats.json（下游出图/出文档唯一数字源）。
当前可从服务商 xlsx 直接复算的字段由本脚本读取；报告确认值在代码中明确保留为 report-only。
用法： .venv/bin/python src/hy7_etl.py
"""
import os, json, warnings
import openpyxl
warnings.filterwarnings("ignore")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
B = os.path.join(ROOT, "data", "hy7_raw")
# 已不再依赖派生汇总 xlsx；只从服务商原始 xlsx 读取。
OUT = os.path.join(ROOT, "experiments", "hy7_stats.json")

# 6 个成像尺度元数据（来自已核验的「数据总览」）
SCALES = [
    {"key":"amics","name":"Amics 矿物扫描","res":"25μm / 1μm","modality":"自动矿物学(BSE+矿物图)",
     "dim":"TIF 图像","content":"矿物组成、元素含量","porosity":None},
    {"key":"ct14","name":"微米CT 柱塞","res":"14μm","modality":"微米CT",
     "dim":"2024³(原始) → 1620×1620×1600(处理)","content":"基质(SUS)/裂缝(FENG)/孔隙(pore) 三维分割","porosity":1.558},
    {"key":"ct28","name":"微米CT 精细","res":"2.8μm","modality":"微米CT",
     "dim":"2024³(原始) → 1500 切片(处理)","content":"基质(SUS)/孔隙(pore) 精细三维","porosity":7.182},
    {"key":"nano65","name":"纳米CT","res":"65nm","modality":"纳米CT",
     "dim":"790 切片序列","content":"纳米级孔隙/喉道/裂缝、逐切片面孔率","porosity":1.925},
    {"key":"maps10","name":"Maps SEM","res":"200nm / 10nm","modality":"扫描电镜拼图",
     "dim":"65600×45000(整体) / 131200×82000(精细)","content":"二维有机/无机孔隙-裂缝分类","porosity":0.411},
    {"key":"fib","name":"FIB-SEM","res":"8.589nm","modality":"聚焦离子束-SEM",
     "dim":"23.19×13.40×7.05 μm, 100+ 切片","content":"三维纳米孔隙-喉道网络拓扑","porosity":None,"organic_pct":1.52},
]
# 各尺度孔喉统计源文件
SRC = {
    "ct14":  f"{B}/微米CT柱塞扫描-14um/孔隙喉道、裂缝统计数据.xlsx",
    "ct28":  f"{B}/微米CT精细扫描-2p8um/孔隙喉道、裂缝统计数据.xlsx",
    "nano65":f"{B}/纳米CT扫描-65nm/孔隙喉道、裂缝统计数据.xlsx",
    "fib":   f"{B}/FIB-SEM聚焦离子束扫描-8p589nm/孔隙喉道、裂缝统计数据.xlsx",
}
MINERAL_SRC = {
    "coarse": f"{B}/Amics矿物整体扫描25um+精细扫描1um/粗扫矿物元素含量.xlsx",
    "fine":   f"{B}/Amics矿物整体扫描25um+精细扫描1um/精扫矿物元素含量.xlsx",
}
MAPS_SRC = f"{B}/Maps整体扫描200nm+精细扫描10nm/孔隙、裂缝计算结果.xlsx"

# These fields are direct service-provider report values, not quantities that
# can be recovered from the staged distribution XLSX alone. The large DOCX
# files remain on Linux; paths/SHA/table identifiers are the reproducible link.
MAPS_REPORT = {
    "remote_path": "/home/user/HXL/HY7_source/吉林大学数据报告归总/Maps整体扫描200nm+精细扫描10nm/花页7井_4199.21mMaps精扫实验报告.docx",
    "sha256": "6202d087a3fe0d80c27d1e71372961fa293b03730837c776eae034168eab9560",
    "table": "表 3-1 样品Maps计算结果",
    "page": None,
    "locator_note": "DOCX 未提供稳定页码；以已哈希文档内的表号定位，渲染版页码不可作为固定 provenance。",
    "total_porosity_pct": 0.4112,
    "classification": [
        {"type": "有机质孔隙", "area_pct": 0.0009, "share_pct": 0.23, "radius_nm": 54.46},
        {"type": "有机质裂缝", "area_pct": 0.0005, "share_pct": 0.12, "radius_nm": 66.56},
        {"type": "无机质孔隙", "area_pct": 0.2893, "share_pct": 70.36, "radius_nm": 63.23},
        {"type": "无机质裂缝", "area_pct": 0.1205, "share_pct": 29.29, "radius_nm": 96.59},
        {"type": "合计", "area_pct": 0.4112, "share_pct": 100.0, "radius_nm": None},
    ],
}
FIB_REPORT = {
    "remote_path": "/home/user/HXL/HY7_source/吉林大学数据报告归总/FIB-SEM聚焦离子束扫描-8p589nm/花页7井-4199.21m-FIB报告.docx",
    "sha256": "03c60566a15f581f0a8e3d9dc4e1fdff7369f7e2a9e46b726d59c79c8d650da6",
    "table": "表 6 花页7井-4199.21m样品孔隙网络型结果",
    "page": None,
    "locator_note": "DOCX 未提供稳定页码；以已哈希文档内的表号定位，渲染版页码不可作为固定 provenance。",
    "voxel_size_nm": [8.589, 8.589, 8.589],
    "organic_pct": 1.52,
    "pores": 76252,
    "throats": 62086,
}

# This is a descriptive aggregation of vendor mineral names, not a formal
# petrographic classification supplied by the vendor. Keep the existing rule
# explicit and versioned so a later domain decision changes it deliberately.
COARSE_MINERAL_GROUPING_V1 = {
    "schema": "hy7-amics-coarse-mineral-groups.v1",
    "evidence_grade": "INTERPRETIVE",
    "purpose": "descriptive mineral aggregation for the coarse Amics table",
    "not_a_vendor_lithology_field": True,
    "groups": {
        "长英质": ["斜长石", "石英", "钾长石"],
        "碳酸盐": ["方解石", "白云石", "铁白云石"],
        "黏土": ["绿泥石", "伊利石", "白云母", "伊蒙混层", "黑云母"],
        "黄铁矿": ["黄铁矿"],
    },
    "remainder_group": "其它",
}
LITHOLOGY_INTERPRETATION = {
    "label": "页岩（钙质长英质）",
    "evidence_grade": "INTERPRETIVE",
    "basis": "COARSE_MINERAL_GROUPING_V1 的描述性长英质与碳酸盐聚合",
    "not_a_vendor_classification": True,
}


def read_block(ws, c_label, c_v1, c_v2=None, start=1):
    """从某列起逐行读 (label, v1[, v2])，遇空 label 停。"""
    out = []
    for r in range(start, ws.max_row):
        lab = ws.cell(row=r+1, column=c_label+1).value
        if lab is None or str(lab).strip() == "":
            if out: break
            continue
        v1 = ws.cell(row=r+1, column=c_v1+1).value
        if not isinstance(v1, (int, float)):
            continue
        rec = {"bin": str(lab).strip(), "v1": round(float(v1), 3)}
        if c_v2 is not None:
            v2 = ws.cell(row=r+1, column=c_v2+1).value
            rec["v2"] = round(float(v2), 3) if isinstance(v2, (int, float)) else None
        out.append(rec)
    return out


def parse_scale(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    res = {}
    if "孔隙数据" in wb.sheetnames:
        ws = wb["孔隙数据"]
        res["pore_radius"] = read_block(ws, 0, 1, 2)        # 半径区间, 体积%, 个数%
        res["coordination"] = read_block(ws, 11, 12, 13)     # 配位数, 体积%, 个数%
    if "喉道数据" in wb.sheetnames:
        ws = wb["喉道数据"]
        res["throat_radius"] = read_block(ws, 0, 1, 2)
        res["throat_length"] = read_block(ws, 11, 12, 13)
    if "裂缝数据" in wb.sheetnames:
        ws = wb["裂缝数据"]
        res["fracture"] = read_block(ws, 0, 1)
    return res


def parse_minerals(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "矿物" not in wb.sheetnames:
        raise ValueError(f"{path}: missing required sheet '矿物'")
    ws = wb["矿物"]
    out = []
    for row in ws.iter_rows(values_only=True):
        if row[1] and isinstance(row[2], (int, float)):
            out.append({"mineral": str(row[1]).strip(), "pct": round(float(row[2]), 2)})
    if not out:
        raise ValueError(f"{path}: sheet '矿物' has no numeric mineral percentage rows")
    return out


def parse_maps(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    res = {}
    for sh in wb.sheetnames:
        ws = wb[sh]; rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 列：半径μm, 半径nm, 分布%
            if len(row) >= 3 and isinstance(row[1], (int, float)) and isinstance(row[2], (int, float)):
                rows.append({"r_nm": round(float(row[1]), 2), "pct": round(float(row[2]), 2)})
        res[sh] = rows
    return res


def infer_total_count(percentages, *, max_total=100_000, tolerance=1e-6):
    """Recover the smallest exact integer denominator from full count percentages.

    The vendor FIB sheets store full-precision bin percentages, not bin counts.
    A candidate total is accepted only when every percentage reconstructs an
    integer bin count and those counts sum to the candidate. The bounded search
    is intentionally fail-closed rather than treating rounded percentages as
    proof of a count.
    """
    values = [float(value) for value in percentages]
    if not values or any(value < 0.0 or value > 100.0 for value in values):
        raise ValueError("count percentages must be a non-empty list in [0, 100]")
    if abs(sum(values) - 100.0) > tolerance:
        raise ValueError("count percentages must sum to 100")
    for total in range(1, max_total + 1):
        counts = [round(value * total / 100.0) for value in values]
        if sum(counts) == total and all(abs(value * total / 100.0 - count) <= tolerance for value, count in zip(values, counts)):
            return total
    raise ValueError(f"no exact count denominator up to {max_total} for the supplied percentages")


def count_percentages(path, sheet):
    """Read the first vendor count-percentage column from a distribution sheet."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"{path}: missing required sheet {sheet!r}")
    rows = list(wb[sheet].iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{path}: sheet {sheet!r} is empty")
    column = next((index for index, value in enumerate(rows[0]) if value and "个数占比" in str(value)), None)
    if column is None:
        raise ValueError(f"{path}: sheet {sheet!r} missing count-percentage column")
    values = [row[column] for row in rows[1:] if len(row) > column and isinstance(row[column], (int, float))]
    if not values:
        raise ValueError(f"{path}: sheet {sheet!r} has no numeric count percentages")
    return values


def fib_counts_from_xlsx(path):
    """Recover FIB pore/throat counts from their complete XLSX distributions."""
    return {
        "pores": infer_total_count(count_percentages(path, "孔隙数据")),
        "throats": infer_total_count(count_percentages(path, "喉道数据")),
    }


def mineral_groups(items, grouping=COARSE_MINERAL_GROUPING_V1):
    """Apply the explicitly versioned descriptive grouping without classifying rock type."""
    groups = {
        name: round(sum(item["pct"] for item in items if item["mineral"] in members), 2)
        for name, members in grouping["groups"].items()
    }
    groups[grouping["remainder_group"]] = round(100 - sum(groups.values()), 2)
    return groups


def facekong_components(scale_dir):
    """Read the vendor face-porosity sheet and preserve its component denominator.

    CT14/nano65 expose fracture and pore percentages separately; CT28 has only
    one total face-porosity column. Returning the component means prevents
    report annotations from redistributing a verified total into unsupported
    pore/fracture values.
    """
    path = f"{B}/{scale_dir}/孔隙喉道、裂缝统计数据.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    sh = next((s for s in wb.sheetnames if "面孔率" in s), None)
    if not sh:
        raise ValueError(f"{path}: missing face-porosity sheet")
    rows = list(wb[sh].iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{path}: face-porosity sheet is empty")
    hdr = rows[0]
    def col(pred):
        for j, h in enumerate(hdr):
            if h and pred(str(h)): return j
        return None
    cf, cp = col(lambda h: "裂缝面孔率" in h), col(lambda h: "孔隙面孔率" in h)
    if cf is not None and cp is not None:        # 14μm / 65nm：裂缝+孔隙分列
        fr = [r[cf] for r in rows[1:] if isinstance(r[cf], (int, float))]
        po = [r[cp] for r in rows[1:] if isinstance(r[cp], (int, float))]
        if not fr or not po:
            raise ValueError(f"{path}: face-porosity sheet has no numeric fracture/pore values")
        if len(fr) != len(po):
            raise ValueError(f"{path}: fracture/pore face-porosity row counts differ ({len(fr)} vs {len(po)})")
        fracture_pct = sum(fr) / len(fr)
        pore_pct = sum(po) / len(po)
        return {
            "sheet": sh,
            "n_slices": len(fr),
            "fracture_pct": round(fracture_pct, 6),
            "pore_pct": round(pore_pct, 6),
            "total_pct": round(fracture_pct + pore_pct, 3),
        }
    ct = col(lambda h: "面孔率" in h and "平均" not in h and "切片" not in h)  # 2.8μm 单列
    if ct is None:
        raise ValueError(f"{path}: face-porosity sheet has no supported porosity column")
    vals = [r[ct] for r in rows[1:] if isinstance(r[ct], (int, float))]
    if not vals:
        raise ValueError(f"{path}: face-porosity sheet has no numeric porosity values")
    return {"sheet": sh, "n_slices": len(vals), "total_pct": round(sum(vals) / len(vals), 3)}


def facekong_total(scale_dir):
    """Backward-compatible total face porosity for callers that need only the rounded total."""
    return facekong_components(scale_dir)["total_pct"]


def parse_porosity_summary():
    """只依赖源数据：CT 三尺度逐切片重算总孔隙度；Maps/FIB 用报告确认值（FIB 1.52% 为有机质占比，非孔隙度）。"""
    ct14 = facekong_components("微米CT柱塞扫描-14um")
    ct28 = facekong_components("微米CT精细扫描-2p8um")
    nano65 = facekong_components("纳米CT扫描-65nm")
    poro = [
        {"scale":"微米CT柱塞扫描","res":"14μm","porosity":ct14["total_pct"],
         "source":f"{ct14['n_slices']} 切片面孔率均值（源文件重算）",
         "note":f"裂缝{ct14['fracture_pct']:.3f}% + 孔隙{ct14['pore_pct']:.3f}%"},
        {"scale":"微米CT精细扫描","res":"2.8μm","porosity":ct28["total_pct"],
         "source":f"{ct28['n_slices']} 切片面孔率均值（源文件重算）","note":"单列面孔率（服务商处理体无 FENG 组件）"},
        {"scale":"纳米CT扫描","res":"65nm","porosity":nano65["total_pct"],
         "source":f"{nano65['n_slices']} 切片面孔率均值（源文件重算）",
         "note":f"裂缝{nano65['fracture_pct']:.3f}% + 孔隙{nano65['pore_pct']:.3f}%"},
        # REPORT_DIRECT: Maps total/class values come from MAPS_REPORT table 3-1.
        {"scale":"Maps SEM精扫","res":"10nm","porosity":round(MAPS_REPORT["total_porosity_pct"], 3),
         "source":"二维面积率统计（Maps 报告）","note":"有机+无机孔隙裂缝总计（2D 面孔率）"},
        {"scale":"FIB-SEM","res":"8.589nm","porosity":FIB_REPORT["organic_pct"],
         "source":"FIB-SEM 报告","note":"有机质占比，非孔隙度"},
    ]
    return poro, MAPS_REPORT["classification"]


def main():
    stats = {
        "well": "花页7", "depth_m": 4199.21, "lithology": LITHOLOGY_INTERPRETATION["label"],
        "source": "吉林大学样品 / 欧勒姆能源测试",
        "scales": SCALES,
        "verified": "数字直接读自各尺度服务商原始统计源文件；总孔隙度由逐切片面孔率独立复现（src/verify_hy7.py 曾12项全过）",
    }
    # 矿物
    coarse = parse_minerals(MINERAL_SRC["coarse"])
    fine = parse_minerals(MINERAL_SRC["fine"])
    stats["minerals"] = {"coarse_25um": coarse, "fine_1um": fine}
    stats["lithology_groups"] = mineral_groups(coarse)
    stats["lithology_grouping_provenance"] = COARSE_MINERAL_GROUPING_V1
    stats["lithology_interpretation"] = LITHOLOGY_INTERPRETATION
    # 各尺度分布
    stats["scale_data"] = {k: parse_scale(p) for k, p in SRC.items()}
    stats["maps_dist"] = parse_maps(MAPS_SRC)
    # 孔隙度（沿用原始报告术语“总孔隙度”；FIB-SEM 的 1.52% 是“有机质占比”非孔隙度）
    poro, classif = parse_porosity_summary()
    METRIC = {"微米CT柱塞扫描":"总孔隙度", "微米CT精细扫描":"总孔隙度", "纳米CT扫描":"总孔隙度",
              "Maps SEM精扫":"总孔隙度(2D面孔率)", "FIB-SEM":"有机质占比"}
    for p in poro:
        p["metric"] = METRIC.get(p["scale"].strip(), "总孔隙度")
    stats["porosity_summary"] = poro
    # 仅“总孔隙度”尺度参与多尺度孔隙度对比（剔除 FIB-SEM 的有机质占比）
    stats["porosity_total"] = [p for p in poro if p["metric"].startswith("总孔隙度")]
    fib = [p for p in poro if p["scale"].strip() == "FIB-SEM"]
    stats["fib_organic_pct"] = fib[0]["porosity"] if fib else 1.52
    stats["maps_classification"] = classif
    # 逐切片重算（佐证，等于原报告“总孔隙度”）
    stats["porosity_rederived"] = {
        "ct14": next(p["porosity"] for p in poro if p["res"] == "14μm"),
        "ct28": next(p["porosity"] for p in poro if p["res"] == "2.8μm"),
        "nano65": next(p["porosity"] for p in poro if p["res"] == "65nm"),
    }
    stats["report_direct_provenance"] = {"maps": MAPS_REPORT, "fib": FIB_REPORT}
    fib_counts = fib_counts_from_xlsx(SRC["fib"])
    if fib_counts != {"pores": FIB_REPORT["pores"], "throats": FIB_REPORT["throats"]}:
        raise ValueError("FIB XLSX-recovered counts disagree with the hashed provider report")
    stats["fib_counts"] = {
        **fib_counts,
        "source": "XLSX_RECOMPUTED from full-precision count-percentage distributions",
        "report_crosscheck": "matches FIB_REPORT table 6",
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    json.dump(stats, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(">> 已生成", os.path.relpath(OUT, ROOT))
    print("   矿物(粗扫)", len(coarse), "种; 岩性组:", stats["lithology_groups"])
    print("   多尺度孔隙度:", [(p["scale"][:6], p["porosity"]) for p in poro])
    for k in SRC:
        sd = stats["scale_data"][k]
        print(f"   {k}: 孔径{len(sd.get('pore_radius',[]))}档 喉径{len(sd.get('throat_radius',[]))}档 "
              f"配位{len(sd.get('coordination',[]))} 裂缝{len(sd.get('fracture',[]))}")


if __name__ == "__main__":
    main()
